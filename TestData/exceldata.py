import openpyxl

Dict={}

book=openpyxl.load_workbook("C:\\Users\\bhavn\\Documents\\pythondemo.xlsx")

sheet=book.active

cell=sheet.cell(row=1,column=1)
print(cell.value)
sheet.cell(row=2 , column=3).value ="Sharma"
cell=sheet.cell(row=2 , column=3)
print(cell.value)
print(sheet.max_row)
print(sheet.max_column)

for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value == "Testcase2":

      for j in range(2,sheet.max_column+1):
       #Dict["lastname"]=sharma
       Dict[sheet.cell(row=1 , column=j).value] =sheet.cell(row=i,column=j).value

print(Dict)

